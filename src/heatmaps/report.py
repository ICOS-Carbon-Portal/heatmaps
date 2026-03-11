# Standard imports
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

# Related imports
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import pandas as pd

# Local imports
from heatmaps.generator import (
    BIN_SETTINGS,
    DOMAIN_FILE_PREFIXES,
    build_heatmap_dataset,
    fetch_raw_data,
    generate_heatmap,
    render_heatmap,
)


def _noop_progress(_: str) -> None:
    """Ignore report progress updates when no listener is provided."""


@dataclass(frozen=True, slots=True)
class ReportResult:
    standalone_paths: list[Path]
    cumulative_paths: list[Path]
    table_paths: list[Path]


def generate_report(
    year: int,
    output_root: Path,
    progress: Callable[[str], None] | None = None,
    cache_dir: Path | None = None,
) -> ReportResult:
    if progress is None:
        progress = _noop_progress

    report_root = output_root / f"report-{year}"
    standalone_dir = report_root / "standalone-years"
    cumulative_dir = report_root / "cumulative"
    tables_dir = report_root / "tables"

    standalone_dir.mkdir(parents=True, exist_ok=True)
    cumulative_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)

    standalone_paths: list[Path] = []
    cumulative_paths: list[Path] = []
    table_paths: list[Path] = []

    for domain in ("atmosphere", "ecosystem"):
        progress(f"Fetching {domain} data")
        raw_data = fetch_raw_data(domain, cache_dir)
        for yr in range(2020, year + 1):
            year_dir = standalone_dir / str(yr)
            for bin_size in ("monthly", "weekly"):
                progress(f"Building {domain} {bin_size} {yr} heatmap")
                standalone_paths.append(
                    generate_heatmap(
                        domain=domain,
                        year=yr,
                        bin_size=bin_size,
                        output_dir=year_dir,
                        raw_data=raw_data,
                    )
                )
        progress(f"Building {domain} monthly cumulative heatmap")
        cumulative_paths.append(
            generate_cumulative_heatmap(
                domain=domain,
                year=year,
                output_dir=cumulative_dir,
                raw_data=raw_data,
            )
        )

        progress(f"Writing {domain} yearly percentages workbook")
        table_paths.append(
            generate_yearly_percentages_workbook(
                domain=domain,
                report_year=year,
                output_dir=tables_dir,
                raw_data=raw_data,
            )
        )

    return ReportResult(
        standalone_paths=standalone_paths,
        cumulative_paths=cumulative_paths,
        table_paths=table_paths,
    )


def generate_cumulative_heatmap(
    domain: str,
    year: int,
    output_dir: Path,
    raw_data: pd.DataFrame | None = None,
) -> Path:
    source_data = raw_data if raw_data is not None else fetch_raw_data(domain)
    start = pd.Timestamp("2020-01-01", tz="UTC")
    end = pd.Timestamp(f"{year}-12-31 23:59:59", tz="UTC")
    parsed_data, percentages = build_heatmap_dataset(
        raw_data=source_data,
        start=start,
        end=end,
        bin_size="monthly",
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = BIN_SETTINGS["monthly"]["suffix"]
    output_path = output_dir / f"heatmap_{domain}_{suffix}_2020_to_{year}.png"
    fig = render_heatmap(
        parsed_data=parsed_data,
        percentages=percentages,
        domain=domain,
        year_label=f"2020 to {year}",
        bin_size="monthly",
    )
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return output_path


def generate_yearly_percentages_workbook(
    domain: str,
    report_year: int,
    output_dir: Path,
    raw_data: pd.DataFrame | None = None,
) -> Path:
    source_data = raw_data if raw_data is not None else fetch_raw_data(domain)
    years = list(range(2020, report_year + 1))
    yearly_label_maps: list[dict[str, str]] = []

    for current_year in years:
        start = pd.Timestamp(f"{current_year}-01-01", tz="UTC")
        end = pd.Timestamp(f"{current_year}-12-31 23:59:59", tz="UTC")
        parsed_data, labels = build_heatmap_dataset(
            raw_data=source_data,
            start=start,
            end=end,
            bin_size="monthly",
        )
        yearly_label_maps.append(
            dict(zip(parsed_data.index.to_list(), labels, strict=True))
        )

    stations = sorted(source_data["station"].unique())
    rows: list[list[str | int]] = []
    for index, station in enumerate(stations, start=1):
        station_rows: list[str | int] = [index, station]
        for label_map in yearly_label_maps:
            station_rows.append(label_map.get(station, "  No Data"))
        rows.append(station_rows)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{DOMAIN_FILE_PREFIXES[domain]}_yearly_percentages.xlsx"
    write_yearly_percentages_workbook(
        output_path=output_path,
        years=years,
        rows=rows,
    )
    return output_path


def write_yearly_percentages_workbook(
    output_path: Path,
    years: list[int],
    rows: list[list[str | int]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    subheader_fill = PatternFill("solid", fgColor="EAF4E3")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    last_year_column = 2 + len(years)
    last_column_letter = get_column_letter(last_year_column)

    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:B2")
    worksheet.merge_cells(f"C1:{last_column_letter}1")

    worksheet["A1"] = ""
    worksheet["B1"] = "Stations"
    worksheet["C1"] = "Yearly percentages"

    for column, year in enumerate(years, start=3):
        worksheet.cell(row=2, column=column, value=year)

    for row_index, row_values in enumerate(rows, start=3):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    for cell in ("B1", "C1"):
        worksheet[cell].fill = header_fill
        worksheet[cell].font = bold
        worksheet[cell].alignment = center
        worksheet[cell].border = border

    for row in (1, 2):
        for column in range(1, last_year_column + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.border = border
            cell.alignment = center
            if row == 2 or column <= 2:
                cell.fill = subheader_fill
                cell.font = bold

    for row in range(3, len(rows) + 3):
        for column in range(1, last_year_column + 1):
            worksheet.cell(row=row, column=column).border = border

    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 16
    for column in range(3, last_year_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 14

    workbook.save(output_path)
